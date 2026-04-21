import argparse
import ast
import csv
import json
import os
import re
import sys
import textwrap
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple

from gemini_secret_store import resolve_api_key_with_source


def bootstrap_openpyxl() -> None:
    try:
        import openpyxl  # noqa: F401
        return
    except ImportError:
        pass

    candidates = [
        Path(__file__).resolve().parent / "venv" / "Lib" / "site-packages",
        Path.cwd() / "venv" / "Lib" / "site-packages",
    ]

    for candidate in candidates:
        if candidate.is_dir():
            sys.path.insert(0, str(candidate))
            try:
                import openpyxl  # noqa: F401
                return
            except ImportError:
                continue

    raise ImportError(
        "openpyxl is required. Install it or use the repo's local venv folder."
    )


bootstrap_openpyxl()

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side
from openpyxl.styles.borders import Border
from openpyxl.utils import get_column_letter


LEVELS = ["Very Low", "Low", "Medium", "High", "Very High"]
LEVEL_TO_SCORE = {level: index for index, level in enumerate(LEVELS)}
MAX_LEVEL_SCORE = LEVEL_TO_SCORE["Very High"]
TARGET_LEVEL_DEFAULT = "High"
PRIORITY_DEFAULT = "Medium"

ASSESSMENT_MODE_QUESTIONNAIRE = "questionnaire"
ASSESSMENT_MODE_FORMAL_EVIDENCE = "formal_evidence"
ASSESSMENT_MODE_LABELS = {
    ASSESSMENT_MODE_QUESTIONNAIRE: "Strict Questionnaire Review",
    ASSESSMENT_MODE_FORMAL_EVIDENCE: "Formal Evidence Review",
}

PRIORITY_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
LEVEL_FILLS = {
    "Very Low": "FDE2E1",
    "Low": "F9E2AF",
    "Medium": "FFF2CC",
    "High": "D9EAD3",
    "Very High": "CFE2F3",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

STOPWORDS = {
    "about",
    "after",
    "again",
    "against",
    "aligned",
    "also",
    "among",
    "and",
    "answer",
    "answers",
    "applicable",
    "assessment",
    "because",
    "before",
    "between",
    "business",
    "company",
    "control",
    "controls",
    "data",
    "describe",
    "document",
    "does",
    "evidence",
    "for",
    "from",
    "have",
    "high",
    "how",
    "information",
    "into",
    "iso",
    "level",
    "managed",
    "management",
    "more",
    "must",
    "needed",
    "none",
    "not",
    "organization",
    "organizational",
    "over",
    "policy",
    "process",
    "provide",
    "question",
    "requested",
    "review",
    "risk",
    "scope",
    "security",
    "should",
    "show",
    "supports",
    "that",
    "the",
    "their",
    "there",
    "these",
    "this",
    "through",
    "under",
    "used",
    "using",
    "what",
    "when",
    "where",
    "which",
    "with",
    "within",
    "your",
}

RUBRIC_ROWS = [
    {
        "Security Level": "Very Low",
        "Numeric Score": 0,
        "Meaning": "Ad hoc, reactive, or absent control with little or no direct evidence.",
        "Expected Evidence": "No formal documentation, inconsistent answers, or missing proof.",
        "Typical Gap": "Control is not established, is largely manual, or exists only informally.",
        "Remediation Focus": "Stand up a minimum viable policy, owner, and baseline process.",
    },
    {
        "Security Level": "Low",
        "Numeric Score": 1,
        "Meaning": "Some activity exists, but execution is inconsistent, narrow, or weakly evidenced.",
        "Expected Evidence": "Partial procedures, isolated examples, or undocumented practices.",
        "Typical Gap": "Coverage is incomplete and the control depends on individuals rather than a system.",
        "Remediation Focus": "Document the control, define scope, and make execution repeatable.",
    },
    {
        "Security Level": "Medium",
        "Numeric Score": 2,
        "Meaning": "Control is defined and partially implemented for key systems, but important gaps remain.",
        "Expected Evidence": "Policies, tickets, reviews, or logs exist for core processes but not for all cases.",
        "Typical Gap": "Control works in major areas but lacks consistency, monitoring, or full evidence.",
        "Remediation Focus": "Close coverage gaps, add review cadences, and strengthen retained evidence.",
    },
    {
        "Security Level": "High",
        "Numeric Score": 3,
        "Meaning": "Documented, consistently implemented, and supported by credible evidence and oversight.",
        "Expected Evidence": "Current standards, records, assigned owners, and evidence of periodic review.",
        "Typical Gap": "Mostly strong, but improvement opportunities remain in automation or metrics.",
        "Remediation Focus": "Improve measurement, efficiency, and cross-functional consistency.",
    },
    {
        "Security Level": "Very High",
        "Numeric Score": 4,
        "Meaning": "Well-governed, measured, and continuously improved across the organization.",
        "Expected Evidence": "Metrics, testing, executive oversight, recurring validation, and strong evidence trails.",
        "Typical Gap": "Limited; focus shifts to optimization and resilience at scale.",
        "Remediation Focus": "Maintain continuous improvement and independent validation.",
    },
]

TEMPLATE_HEADERS = [
    "Section",
    "Subsection",
    "Control_ID",
    "Control_Name",
    "Question_Text",
    "Answer_Type",
    "Answer_Options",
    "Evidence_Requested",
    "Standard_Mapping",
    "Notes",
    "Target_Security_Level",
    "Weakness_Priority",
    "Very_Low_Rubric",
    "Low_Rubric",
    "Medium_Rubric",
    "High_Rubric",
    "Very_High_Rubric",
]
FILLED_TEMPLATE_HEADERS = TEMPLATE_HEADERS + [
    "Current_Security_Level",
    "Current_Level_Score",
    "Meets_Target",
    "Priority",
    "Confidence",
    "Answer_Summary",
    "Evidence_Used",
    "Gap_Summary",
    "Risk_Impact",
    "Remediation_Steps",
    "Rationale",
]
ASSESSMENT_HEADERS = [
    "Section",
    "Subsection",
    "Control_ID",
    "Control_Name",
    "Question_Text",
    "Evidence_Requested",
    "Target_Security_Level",
    "Security_Level",
    "Level_Score",
    "Answer_Summary",
    "Evidence_Used",
    "Gap_Summary",
    "Risk_Impact",
    "Remediation_Steps",
    "Priority",
    "Confidence",
    "Rationale",
]


@dataclass
class Control:
    section: str
    subsection: str
    control_id: str
    control_name: str
    question_text: str
    answer_type: str
    answer_options: str
    evidence_requested: str
    criticality_level: int
    standard_mapping: str
    notes: str
    target_security_level: str
    weakness_priority: str


@dataclass
class EvidenceChunk:
    source: str
    text: str
    tokens: set[str]
    document_path: str
    document_type: str


@dataclass
class ControlEvaluation:
    control_id: str
    answer_summary: str
    evidence_used: List[str] = field(default_factory=list)
    security_level: str = "Very Low"
    level_score: int = 0
    gap_summary: str = ""
    risk_impact: str = ""
    remediation_steps: List[str] = field(default_factory=list)
    priority: str = "High"
    confidence: str = "Low"
    rationale: str = ""


def collapse_whitespace(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_level(value: Any) -> str:
    candidate = collapse_whitespace(value).lower()
    for level in LEVELS:
        if candidate == level.lower():
            return level
    return "Very Low"


def normalize_priority(value: Any) -> str:
    candidate = collapse_whitespace(value).lower()
    for priority in PRIORITY_ORDER:
        if candidate == priority.lower():
            return priority
    return "Medium"


def normalize_confidence(value: Any) -> str:
    candidate = collapse_whitespace(value).lower()
    if candidate in {"high", "medium", "low"}:
        return candidate.title()
    return "Low"


def score_to_level(percent: float) -> str:
    if percent < 20:
        return "Very Low"
    if percent < 40:
        return "Low"
    if percent < 60:
        return "Medium"
    if percent < 80:
        return "High"
    return "Very High"


def percent_to_letter_grade(percent: float) -> str:
    if percent >= 90:
        return "A"
    if percent >= 80:
        return "B"
    if percent >= 70:
        return "C"
    if percent >= 60:
        return "D"
    return "F"


def parse_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default
    text = collapse_whitespace(value)
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def tokenize(text: str) -> set[str]:
    tokens = {
        token.lower()
        for token in re.findall(r"[A-Za-z0-9][A-Za-z0-9_\-/]+", text)
        if len(token) > 2
    }
    return {token for token in tokens if token not in STOPWORDS}


def control_keywords(control: Control) -> set[str]:
    combined = " ".join(
        [
            control.section,
            control.subsection,
            control.control_id,
            control.control_name,
            control.question_text,
            control.evidence_requested,
            control.standard_mapping,
        ]
    )
    return tokenize(combined)


def required_template_columns() -> List[Tuple[str, ...]]:
    return [
        ("Section",),
        ("Subsection",),
        ("Control_ID",),
        ("Control_Name",),
        ("Question_Text",),
        ("Answer_Type",),
        ("Answer_Options",),
        ("Evidence_Requested",),
        ("Standard_Mapping",),
        ("Notes",),
    ]


def resolve_header(headers: Sequence[str], *candidates: str) -> Optional[str]:
    for candidate in candidates:
        if candidate in headers:
            return candidate
    return None


def read_template_controls(template_path: str) -> List[Control]:
    workbook = load_workbook(template_path, data_only=True, read_only=True)
    sheet = workbook["Template"] if "Template" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [collapse_whitespace(value) for value in next(rows)]

    for column_group in required_template_columns():
        if not resolve_header(headers, *column_group):
            raise ValueError(f"Missing template column: {' or '.join(column_group)}")
    legacy_weight_header = resolve_header(headers, "Criticality_Level", "Control_Weight", "Weight")

    controls: List[Control] = []
    for row in rows:
        row_values = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
        control_id = collapse_whitespace(row_values.get("Control_ID"))
        if not control_id:
            continue

        criticality = parse_int(row_values.get(legacy_weight_header), default=0) if legacy_weight_header else 0
        target_level = normalize_level(row_values.get("Target_Security_Level")) if "Target_Security_Level" in headers else TARGET_LEVEL_DEFAULT
        weakness_priority = normalize_priority(row_values.get("Weakness_Priority")) if "Weakness_Priority" in headers else PRIORITY_DEFAULT

        controls.append(
            Control(
                section=collapse_whitespace(row_values.get("Section")),
                subsection=collapse_whitespace(row_values.get("Subsection")),
                control_id=control_id,
                control_name=collapse_whitespace(row_values.get("Control_Name")),
                question_text=collapse_whitespace(row_values.get("Question_Text")),
                answer_type=collapse_whitespace(row_values.get("Answer_Type")),
                answer_options=collapse_whitespace(row_values.get("Answer_Options")),
                evidence_requested=collapse_whitespace(row_values.get("Evidence_Requested")),
                criticality_level=criticality,
                standard_mapping=collapse_whitespace(row_values.get("Standard_Mapping")),
                notes=collapse_whitespace(row_values.get("Notes")),
                target_security_level=target_level,
                weakness_priority=weakness_priority,
            )
        )

    if not controls:
        raise ValueError("No controls found in the template workbook.")

    return controls


def load_questionnaire_chunks(paths: Sequence[str]) -> List[EvidenceChunk]:
    chunks: List[EvidenceChunk] = []

    for path_str in paths:
        path = Path(path_str)
        if not path.exists():
            print(f"[WARN] File not found, skipping: {path}")
            continue

        ext = path.suffix.lower()

        try:
            if ext in {".xlsx", ".xlsm"}:
                workbook = load_workbook(path, data_only=True, read_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        values = [collapse_whitespace(cell) for cell in row if collapse_whitespace(cell)]
                        if not values:
                            continue
                        text = " | ".join(values)
                        if len(text) < 8:
                            continue
                        chunks.append(
                            EvidenceChunk(
                                source=f"{path.name} | {sheet_name} | row {row_index}",
                                text=text,
                                tokens=tokenize(text),
                                document_path=str(path),
                                document_type="Excel",
                            )
                        )
            elif ext == ".csv":
                with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
                    reader = csv.reader(handle)
                    for row_index, row in enumerate(reader, start=1):
                        values = [collapse_whitespace(cell) for cell in row if collapse_whitespace(cell)]
                        if not values:
                            continue
                        text = " | ".join(values)
                        chunks.append(
                            EvidenceChunk(
                                source=f"{path.name} | row {row_index}",
                                text=text,
                                tokens=tokenize(text),
                                document_path=str(path),
                                document_type="CSV",
                            )
                        )
            elif ext in {".txt", ".md"}:
                content = path.read_text(encoding="utf-8", errors="ignore")
                paragraphs = [collapse_whitespace(part) for part in re.split(r"\n\s*\n", content)]
                for index, paragraph in enumerate(paragraphs, start=1):
                    if len(paragraph) < 8:
                        continue
                    chunks.append(
                        EvidenceChunk(
                            source=f"{path.name} | paragraph {index}",
                            text=paragraph,
                            tokens=tokenize(paragraph),
                            document_path=str(path),
                            document_type="Text",
                        )
                    )
            else:
                print(f"[WARN] Unsupported file type, skipping: {path}")
        except Exception as exc:
            print(f"[WARN] Failed to load {path}: {exc}")

    if not chunks:
        raise ValueError("No questionnaire content could be loaded from the provided files.")

    return chunks


def select_relevant_evidence(
    controls: Sequence[Control],
    chunks: Sequence[EvidenceChunk],
    max_chunks: int,
    top_chunks_per_control: int,
    max_chars: int,
) -> str:
    selected_scores: Dict[int, int] = {}

    for control in controls:
        keywords = control_keywords(control)
        ranked: List[Tuple[int, int]] = []
        for index, chunk in enumerate(chunks):
            overlap = len(keywords & chunk.tokens)
            if overlap > 0:
                ranked.append((overlap, index))
        ranked.sort(key=lambda item: (-item[0], len(chunks[item[1]].text)))
        for score, index in ranked[:top_chunks_per_control]:
            selected_scores[index] = max(selected_scores.get(index, 0), score)

    if not selected_scores:
        for index in range(min(max_chunks, len(chunks))):
            selected_scores[index] = 0

    ordered_indices = sorted(
        selected_scores,
        key=lambda index: (-selected_scores[index], len(chunks[index].text), chunks[index].source),
    )

    selected_blocks: List[str] = []
    current_length = 0
    for index in ordered_indices:
        chunk = chunks[index]
        block = f"[Source: {chunk.source}] {chunk.text}"
        if selected_blocks and current_length + len(block) > max_chars:
            break
        selected_blocks.append(block)
        current_length += len(block)
        if len(selected_blocks) >= max_chunks:
            break

    return "\n".join(selected_blocks)


def extract_json_object(text: str) -> Dict[str, Any]:
    raw_text = str(text)
    cleaned = raw_text.strip()
    if not cleaned:
        raise ValueError("Gemini returned an empty response.")

    if cleaned.startswith("```") and cleaned.endswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*```$", "", cleaned)

    def attempt_parse(candidate: str) -> Optional[Dict[str, Any]]:
        try:
            parsed = json.loads(candidate)
            return parsed if isinstance(parsed, dict) else None
        except json.JSONDecodeError:
            return None

    def attempt_python_literal(candidate: str) -> Optional[Dict[str, Any]]:
        try:
            parsed = ast.literal_eval(candidate)
            return parsed if isinstance(parsed, dict) else None
        except (ValueError, SyntaxError):
            return None

    def pythonize_json_literals(candidate: str) -> str:
        candidate = re.sub(r"\btrue\b", "True", candidate)
        candidate = re.sub(r"\bfalse\b", "False", candidate)
        candidate = re.sub(r"\bnull\b", "None", candidate)
        return candidate

    def normalize_quotes(candidate: str) -> str:
        return (
            candidate.replace("\u201c", '"')
            .replace("\u201d", '"')
            .replace("\u2018", "'")
            .replace("\u2019", "'")
        )

    parsed = attempt_parse(cleaned)
    if parsed is not None:
        return parsed

    parsed = attempt_python_literal(cleaned)
    if parsed is not None:
        return parsed

    parsed = attempt_python_literal(pythonize_json_literals(cleaned))
    if parsed is not None:
        return parsed

    start = cleaned.find("{")
    end = cleaned.rfind("}")
    if start != -1 and end != -1 and end > start:
        sliced = cleaned[start:end + 1]
        parsed = attempt_parse(sliced)
        if parsed is not None:
            return parsed
        parsed = attempt_python_literal(sliced)
        if parsed is not None:
            return parsed
        parsed = attempt_python_literal(pythonize_json_literals(sliced))
        if parsed is not None:
            return parsed
    else:
        sliced = cleaned

    repaired = normalize_quotes(sliced)
    repaired = re.sub(r"\bTrue\b", "true", repaired)
    repaired = re.sub(r"\bFalse\b", "false", repaired)
    repaired = re.sub(r"\bNone\b", "null", repaired)
    repaired = re.sub(r",\s*([}\]])", r"\1", repaired)
    repaired = re.sub(
        r'([\{,]\s*)([A-Za-z_][A-Za-z0-9_\-./ ]*)(\s*:)',
        lambda match: f'{match.group(1)}"{match.group(2).strip()}"{match.group(3)}',
        repaired,
    )

    parsed = attempt_parse(repaired)
    if parsed is not None:
        return parsed

    parsed = attempt_python_literal(repaired)
    if parsed is not None:
        return parsed

    parsed = attempt_python_literal(pythonize_json_literals(repaired))
    if parsed is not None:
        return parsed

    snippet = collapse_whitespace(raw_text)[:300]
    raise ValueError(f"Gemini response was not valid JSON and could not be repaired. Response snippet: {snippet}")


def normalize_model_name(model_name: str) -> str:
    model_name = collapse_whitespace(model_name)
    return model_name if model_name.startswith("models/") else f"models/{model_name}"


def normalize_assessment_mode(value: Any) -> str:
    candidate = collapse_whitespace(value).lower().replace("-", "_").replace(" ", "_")
    aliases = {
        "questionnaire_review": ASSESSMENT_MODE_QUESTIONNAIRE,
        "strict_questionnaire": ASSESSMENT_MODE_QUESTIONNAIRE,
        "strict_questionnaire_review": ASSESSMENT_MODE_QUESTIONNAIRE,
        "formal": ASSESSMENT_MODE_FORMAL_EVIDENCE,
        "formal_evidence_review": ASSESSMENT_MODE_FORMAL_EVIDENCE,
        "evidence": ASSESSMENT_MODE_FORMAL_EVIDENCE,
        "evidence_review": ASSESSMENT_MODE_FORMAL_EVIDENCE,
    }
    candidate = aliases.get(candidate, candidate)
    if candidate in ASSESSMENT_MODE_LABELS:
        return candidate
    return ASSESSMENT_MODE_QUESTIONNAIRE


def assessment_mode_label(mode: str) -> str:
    return ASSESSMENT_MODE_LABELS[normalize_assessment_mode(mode)]


def build_control_payload(controls: Sequence[Control]) -> List[Dict[str, str]]:
    return [
        {
            "control_id": control.control_id,
            "section": control.section,
            "subsection": control.subsection,
            "control_name": control.control_name,
            "question": control.question_text,
            "evidence_requested": control.evidence_requested,
            "target_security_level": control.target_security_level,
            "weakness_priority_if_below_target": control.weakness_priority,
        }
        for control in controls
    ]


def build_rubric_prompt_text() -> str:
    return "\n".join(
        f"- {row['Security Level']}: {row['Meaning']} Expected evidence: {row['Expected Evidence']}"
        for row in RUBRIC_ROWS
    )


def build_response_contract() -> str:
    return textwrap.dedent(
        """
        Return ONLY valid JSON in this exact structure:
        {
          "controls": {
            "<CONTROL_ID>": {
              "answer_summary": "2-4 sentence evidence-based summary",
              "evidence_used": ["short evidence point with source hint", "short evidence point with source hint"],
              "security_level": "Very Low|Low|Medium|High|Very High",
              "gap_summary": "Explain what is missing or weak",
              "risk_impact": "Explain the likely business or security impact",
              "remediation_steps": ["specific remediation step", "specific remediation step", "specific remediation step"],
              "priority": "Critical|High|Medium|Low",
              "confidence": "High|Medium|Low",
              "rationale": "Explain why the level was assigned"
            }
          }
        }
        """
    ).strip()


def build_questionnaire_assessment_prompt(
    company_name: str,
    controls: Sequence[Control],
    evidence_text: str,
) -> str:
    control_payload = build_control_payload(controls)
    rubric = build_rubric_prompt_text()
    response_contract = build_response_contract()

    return textwrap.dedent(
        f"""
        You are an ISO 27001 / ISO 27002 assessor performing a strict questionnaire-based security review for a smaller company.

        Assessment mode: Strict Questionnaire Review
        Company: "{company_name}"

        Evaluate each control using only the supplied questionnaire excerpts. Be strict, but calibrate for completed security questionnaires:
        - Treat detailed questionnaire answers, named policies or processes, stated owners, review cadences, tool names, scoped control descriptions, certifications, audit-report references, and framework crosswalks as evidence when they directly answer the control.
        - Do not give credit for generic "yes", "in place", marketing language, vague maturity claims, or answers that do not address the control.
        - If an answer is only a bare self-attestation with no scope, cadence, owner, technical detail, or corroborating source, cap the control at Low.
        - Medium requires a defined control with enough detail to understand how it works, but may still lack primary artifacts such as logs, tickets, screenshots, or registers.
        - High requires specific, current, scoped, and consistent questionnaire evidence, preferably corroborated by certification, SOC/ISO/PCI references, crosswalks, named tools, or multiple independent excerpts. Missing primary artifacts should remain a gap.
        - Very High requires evidence of measurement, recurring validation, oversight, continuous improvement, or direct artifact references. Do not award Very High for an ordinary questionnaire attestation alone.
        - If evidence is conflicting, stale, implied, or incomplete, choose the lower reasonable level and explain why.
        - Keep remediation actionable and realistic for a small or mid-sized business.

        Security level rubric:
        {rubric}

        {response_contract}

        Additional rules:
        - Evaluate every control provided.
        - Do not invent evidence that is not in the excerpts.
        - If the questionnaire evidence is enough for partial credit but primary proof is missing, state both facts clearly.
        - Keep remediation steps short, concrete, and implementation-ready.
        - Do not include markdown or any extra keys.

        === CONTROLS TO ASSESS ===
        {json.dumps(control_payload, ensure_ascii=False, indent=2)}

        === EVIDENCE EXCERPTS ===
        {evidence_text}
        """
    ).strip()


def build_formal_evidence_assessment_prompt(
    company_name: str,
    controls: Sequence[Control],
    evidence_text: str,
) -> str:
    control_payload = build_control_payload(controls)
    rubric = build_rubric_prompt_text()
    response_contract = build_response_contract()

    return textwrap.dedent(
        f"""
        You are an ISO 27001 / ISO 27002 assessor performing a formal evidence sufficiency review for a smaller company.

        Assessment mode: Formal Evidence Review
        Company: "{company_name}"

        Evaluate each control using only the supplied evidence excerpts. This mode is stricter than questionnaire review:
        - Treat questionnaire answers as pointers to evidence, not as full proof by themselves.
        - Strong evidence includes policies, standards, procedures, assigned owners, retained records, review minutes, logs, tickets, screenshots, system exports, risk registers, audit reports, test results, diagrams, configuration evidence, or other artifacts showing the control operates.
        - A bare self-attestation or generic "yes" answer should normally be Very Low or Low, even if it claims the control exists.
        - Medium requires credible detail plus at least some artifact-like support or multiple consistent excerpts showing the control is defined and partially implemented.
        - High requires documented, consistently implemented control evidence with credible oversight or review records. Do not award High solely because a questionnaire says the control exists.
        - Very High requires measured operation, recurring validation, executive or governance oversight, and evidence of continuous improvement.
        - Certifications and SOC/ISO/PCI references can corroborate evidence, but they do not automatically satisfy unrelated controls unless the excerpt directly supports that control.
        - If requested artifacts are missing, stale, indirect, or not uploaded, identify the evidence gap and choose the lower reasonable level.
        - Keep remediation actionable and realistic for a small or mid-sized business.

        Security level rubric:
        {rubric}

        {response_contract}

        Additional rules:
        - Evaluate every control provided.
        - Do not invent evidence that is not in the excerpts.
        - If evidence is missing, state that clearly in the summary and gaps.
        - Keep remediation steps short, concrete, and implementation-ready.
        - Do not include markdown or any extra keys.

        === CONTROLS TO ASSESS ===
        {json.dumps(control_payload, ensure_ascii=False, indent=2)}

        === EVIDENCE EXCERPTS ===
        {evidence_text}
        """
    ).strip()


def build_assessment_prompt(
    company_name: str,
    controls: Sequence[Control],
    evidence_text: str,
    assessment_mode: str = ASSESSMENT_MODE_QUESTIONNAIRE,
) -> str:
    mode = normalize_assessment_mode(assessment_mode)
    if mode == ASSESSMENT_MODE_FORMAL_EVIDENCE:
        return build_formal_evidence_assessment_prompt(company_name, controls, evidence_text)
    return build_questionnaire_assessment_prompt(company_name, controls, evidence_text)


def call_gemini_json(api_key: str, model_name: str, prompt: str) -> Dict[str, Any]:
    model_name = normalize_model_name(model_name)
    url = (
        "https://generativelanguage.googleapis.com/v1beta/"
        f"{urllib.parse.quote(model_name, safe='/')}:generateContent?key={urllib.parse.quote(api_key)}"
    )

    payload = {
        "contents": [
            {
                "role": "user",
                "parts": [{"text": prompt}],
            }
        ],
        "generationConfig": {
            "temperature": 0.1,
            "responseMimeType": "application/json",
        },
    }

    request = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urllib.request.urlopen(request, timeout=180) as response:
            raw_response = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"Gemini HTTP error {exc.code}: {detail}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"Gemini request failed: {exc}") from exc

    candidates = raw_response.get("candidates") or []
    if not candidates:
        prompt_feedback = raw_response.get("promptFeedback")
        raise RuntimeError(f"Gemini returned no candidates. Prompt feedback: {prompt_feedback}")

    parts = candidates[0].get("content", {}).get("parts", [])
    response_text = "\n".join(part.get("text", "") for part in parts if part.get("text"))
    parsed = extract_json_object(response_text)

    if not isinstance(parsed, dict):
        raise ValueError("Gemini response was not a JSON object.")

    return parsed


def normalize_text_list(value: Any, fallback: Optional[List[str]] = None) -> List[str]:
    if isinstance(value, list):
        normalized = [collapse_whitespace(item) for item in value if collapse_whitespace(item)]
        if normalized:
            return normalized
    if isinstance(value, str) and collapse_whitespace(value):
        return [collapse_whitespace(value)]
    return fallback or []


def elevate_priority(priority: str, baseline: str, level_score: int) -> str:
    priority = normalize_priority(priority)
    baseline = normalize_priority(baseline)
    chosen = priority if PRIORITY_ORDER[priority] <= PRIORITY_ORDER[baseline] else baseline
    if level_score == 0 and PRIORITY_ORDER[chosen] > PRIORITY_ORDER["High"]:
        return "High"
    return chosen


def build_evaluations_from_response(
    controls: Sequence[Control],
    response_json: Dict[str, Any],
) -> Dict[str, ControlEvaluation]:
    response_controls = response_json.get("controls", {})
    evaluations: Dict[str, ControlEvaluation] = {}

    for control in controls:
        raw = response_controls.get(control.control_id, {}) if isinstance(response_controls, dict) else {}
        level = normalize_level(raw.get("security_level"))
        level_score = LEVEL_TO_SCORE[level]

        answer_summary = collapse_whitespace(raw.get("answer_summary")) or "No direct evidence found in the provided questionnaires."
        evidence_used = normalize_text_list(
            raw.get("evidence_used"),
            fallback=["No direct evidence cited by the model."],
        )
        remediation_steps = normalize_text_list(
            raw.get("remediation_steps"),
            fallback=[
                "Define a documented owner and minimum control baseline.",
                "Collect evidence showing the control is operating.",
                "Review the control against the target maturity level.",
            ],
        )

        priority = elevate_priority(
            raw.get("priority"),
            control.weakness_priority,
            level_score,
        )

        evaluations[control.control_id] = ControlEvaluation(
            control_id=control.control_id,
            answer_summary=answer_summary,
            evidence_used=evidence_used,
            security_level=level,
            level_score=level_score,
            gap_summary=collapse_whitespace(raw.get("gap_summary")) or "Evidence does not yet show the control operating at the target level.",
            risk_impact=collapse_whitespace(raw.get("risk_impact")) or "Unclear or inconsistent control execution may increase the likelihood of security or compliance gaps.",
            remediation_steps=remediation_steps,
            priority=priority,
            confidence=normalize_confidence(raw.get("confidence")),
            rationale=collapse_whitespace(raw.get("rationale")) or "The assigned level reflects the strength and completeness of the available questionnaire evidence.",
        )

    return evaluations


def is_json_response_error(exc: Exception) -> bool:
    message = collapse_whitespace(str(exc)).lower()
    markers = [
        "not valid json",
        "jsondecodeerror",
        "expecting property name enclosed in double quotes",
        "unterminated string",
    ]
    return any(marker in message for marker in markers)


def fallback_evaluation_for_control(control: Control, exc: Exception) -> ControlEvaluation:
    short_error = collapse_whitespace(str(exc))[:220]
    return ControlEvaluation(
        control_id=control.control_id,
        answer_summary=(
            "Automated assessment could not parse a valid structured response for this control. "
            "A conservative fallback rating was applied."
        ),
        evidence_used=[f"Model parse failure: {short_error}"],
        security_level="Very Low",
        level_score=LEVEL_TO_SCORE["Very Low"],
        gap_summary="Structured model output was unavailable, so evidence could not be reliably scored.",
        risk_impact="Without a reliable structured assessment, this control should be treated as below target until reviewed.",
        remediation_steps=[
            "Re-run this control with a smaller prompt or batch size.",
            "Verify the evidence excerpts for this control are concise and relevant.",
            "Manually review this control result before final sign-off.",
        ],
        priority=control.weakness_priority,
        confidence="Low",
        rationale="Fallback rating applied because the model response could not be parsed into valid JSON.",
    )


def assess_control_batch(
    api_key: str,
    model_name: str,
    company_name: str,
    batch_controls: Sequence[Control],
    chunks: Sequence[EvidenceChunk],
    max_chunks: int,
    top_chunks_per_control: int,
    max_evidence_chars: int,
    assessment_mode: str,
    batch_start_index: int,
) -> Dict[str, ControlEvaluation]:
    batch_end_index = batch_start_index + len(batch_controls)
    evidence_text = select_relevant_evidence(
        batch_controls,
        chunks,
        max_chunks=max_chunks,
        top_chunks_per_control=top_chunks_per_control,
        max_chars=max_evidence_chars,
    )
    prompt = build_assessment_prompt(company_name, batch_controls, evidence_text, assessment_mode)

    try:
        response_json = call_gemini_json(api_key, model_name, prompt)
        print(f"[INFO] Assessed controls {batch_start_index + 1}-{batch_end_index}")
        return build_evaluations_from_response(batch_controls, response_json)
    except Exception as exc:
        if not is_json_response_error(exc):
            raise

        if len(batch_controls) > 1:
            split_at = max(1, len(batch_controls) // 2)
            print(
                f"[WARN] JSON parse failed for controls {batch_start_index + 1}-{batch_end_index}; "
                "retrying this batch in smaller pieces."
            )
            left = assess_control_batch(
                api_key=api_key,
                model_name=model_name,
                company_name=company_name,
                batch_controls=batch_controls[:split_at],
                chunks=chunks,
                max_chunks=max_chunks,
                top_chunks_per_control=top_chunks_per_control,
                max_evidence_chars=max_evidence_chars,
                assessment_mode=assessment_mode,
                batch_start_index=batch_start_index,
            )
            right = assess_control_batch(
                api_key=api_key,
                model_name=model_name,
                company_name=company_name,
                batch_controls=batch_controls[split_at:],
                chunks=chunks,
                max_chunks=max_chunks,
                top_chunks_per_control=top_chunks_per_control,
                max_evidence_chars=max_evidence_chars,
                assessment_mode=assessment_mode,
                batch_start_index=batch_start_index + split_at,
            )
            left.update(right)
            return left

        control = batch_controls[0]
        print(
            f"[WARN] JSON parse failed for control {control.control_id}; "
            "using conservative fallback evaluation."
        )
        return {control.control_id: fallback_evaluation_for_control(control, exc)}


def assess_controls(
    api_key: str,
    model_name: str,
    company_name: str,
    controls: Sequence[Control],
    chunks: Sequence[EvidenceChunk],
    batch_size: int,
    max_chunks: int,
    top_chunks_per_control: int,
    max_evidence_chars: int,
    assessment_mode: str = ASSESSMENT_MODE_QUESTIONNAIRE,
    max_parallel_batches: int = 1,
    progress_callback: Optional[Callable[[int, int, int], None]] = None,
) -> Dict[str, ControlEvaluation]:
    all_evaluations: Dict[str, ControlEvaluation] = {}
    assessment_mode = normalize_assessment_mode(assessment_mode)

    total_controls = len(controls)
    normalized_batch_size = max(1, batch_size)
    batches = [
        (start, min(start + normalized_batch_size, total_controls), controls[start:min(start + normalized_batch_size, total_controls)])
        for start in range(0, total_controls, normalized_batch_size)
    ]
    completed_controls = 0
    worker_count = max(1, max_parallel_batches)

    if worker_count == 1 or len(batches) <= 1:
        for start, end, batch_controls in batches:
            batch_evaluations = assess_control_batch(
                api_key=api_key,
                model_name=model_name,
                company_name=company_name,
                batch_controls=batch_controls,
                chunks=chunks,
                max_chunks=max_chunks,
                top_chunks_per_control=top_chunks_per_control,
                max_evidence_chars=max_evidence_chars,
                assessment_mode=assessment_mode,
                batch_start_index=start,
            )
            all_evaluations.update(batch_evaluations)
            completed_controls += end - start
            if progress_callback is not None:
                progress_callback(completed_controls, completed_controls, total_controls)
        return all_evaluations

    with ThreadPoolExecutor(max_workers=min(worker_count, len(batches))) as executor:
        futures = {}
        for start, end, batch_controls in batches:
            future = executor.submit(
                assess_control_batch,
                api_key,
                model_name,
                company_name,
                batch_controls,
                chunks,
                max_chunks,
                top_chunks_per_control,
                max_evidence_chars,
                assessment_mode,
                start,
            )
            futures[future] = (start, end)

        for future in as_completed(futures):
            start, end = futures[future]
            batch_evaluations = future.result()
            all_evaluations.update(batch_evaluations)
            completed_controls += end - start
            print(
                f"[INFO] Completed {completed_controls}/{total_controls} controls "
                f"with {min(worker_count, len(batches))} parallel batch workers."
            )
            if progress_callback is not None:
                progress_callback(completed_controls, completed_controls, total_controls)

    return all_evaluations


def summarize_sections(
    controls: Sequence[Control],
    evaluations: Dict[str, ControlEvaluation],
) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Tuple[Control, ControlEvaluation]]] = defaultdict(list)
    for control in controls:
        grouped[control.section].append((control, evaluations[control.control_id]))

    summaries: List[Dict[str, Any]] = []
    for section, items in grouped.items():
        earned = sum(item[1].level_score for item in items)
        maximum = len(items) * MAX_LEVEL_SCORE
        percent = (earned / maximum) * 100.0 if maximum else 0.0
        below_target = sum(
            1
            for control, evaluation in items
            if evaluation.level_score < LEVEL_TO_SCORE[control.target_security_level]
        )
        high_priority = sum(
            1
            for control, evaluation in items
            if evaluation.level_score < LEVEL_TO_SCORE[control.target_security_level]
            and evaluation.priority in {"Critical", "High"}
        )
        level_counts = Counter(evaluation.security_level for _, evaluation in items)

        summaries.append(
            {
                "Section": section,
                "Control Count": len(items),
                "Level Score Total": round(earned, 2),
                "Level Score Max": round(maximum, 2),
                "Score Percent": round(percent, 2),
                "Security Level": score_to_level(percent),
                "Below Target Controls": below_target,
                "High Priority Findings": high_priority,
                "Very Low": level_counts.get("Very Low", 0),
                "Low": level_counts.get("Low", 0),
                "Medium": level_counts.get("Medium", 0),
                "High": level_counts.get("High", 0),
                "Very High": level_counts.get("Very High", 0),
            }
        )

    summaries.sort(key=lambda row: row["Score Percent"])
    return summaries


def build_findings(
    controls: Sequence[Control],
    evaluations: Dict[str, ControlEvaluation],
) -> List[Dict[str, Any]]:
    findings: List[Dict[str, Any]] = []
    for control in controls:
        evaluation = evaluations[control.control_id]
        target_score = LEVEL_TO_SCORE[control.target_security_level]
        if evaluation.level_score >= target_score:
            continue

        severity = target_score - evaluation.level_score
        findings.append(
            {
                "Priority": evaluation.priority,
                "Severity Score": severity,
                "Section": control.section,
                "Subsection": control.subsection,
                "Control ID": control.control_id,
                "Control Name": control.control_name,
                "Current Level": evaluation.security_level,
                "Target Level": control.target_security_level,
                "Question": control.question_text,
                "Gap Summary": evaluation.gap_summary,
                "Risk Impact": evaluation.risk_impact,
                "Remediation Steps": "\n".join(f"- {step}" for step in evaluation.remediation_steps),
                "Evidence Used": "\n".join(f"- {item}" for item in evaluation.evidence_used),
                "Confidence": evaluation.confidence,
            }
        )

    findings.sort(
        key=lambda row: (
            PRIORITY_ORDER.get(row["Priority"], 99),
            -row["Severity Score"],
            row["Section"],
            row["Control ID"],
        )
    )
    return findings


def overall_metrics(
    controls: Sequence[Control],
    evaluations: Dict[str, ControlEvaluation],
) -> Dict[str, Any]:
    earned = sum(evaluations[control.control_id].level_score for control in controls)
    maximum = len(controls) * MAX_LEVEL_SCORE
    percent = (earned / maximum) * 100.0 if maximum else 0.0
    controls_meeting_target = sum(
        1
        for control in controls
        if evaluations[control.control_id].level_score >= LEVEL_TO_SCORE[control.target_security_level]
    )

    return {
        "level_score_total": round(earned, 2),
        "level_score_max": round(maximum, 2),
        "average_level_score": round((earned / len(controls)) if controls else 0.0, 2),
        "score_percent": round(percent, 2),
        "security_level": score_to_level(percent),
        "letter_grade": percent_to_letter_grade(percent),
        "control_count": len(controls),
        "controls_meeting_target": controls_meeting_target,
        "controls_below_target": len(controls) - controls_meeting_target,
    }


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER


def style_header_row(ws, row_number: int = 1) -> None:
    for cell in ws[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT
        cell.border = THIN_BORDER


def autosize_columns(ws, max_width: int = 60) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), max_width)


def write_rows(ws, headers: Sequence[str], rows: Iterable[Dict[str, Any]]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_header_row(ws)
    format_sheet(ws)
    autosize_columns(ws)


def template_row(control: Control) -> List[Any]:
    return [
        control.section,
        control.subsection,
        control.control_id,
        control.control_name,
        control.question_text,
        control.answer_type,
        control.answer_options,
        control.evidence_requested,
        control.standard_mapping,
        control.notes,
        control.target_security_level,
        control.weakness_priority,
        control_rubric_text(control, "Very Low"),
        control_rubric_text(control, "Low"),
        control_rubric_text(control, "Medium"),
        control_rubric_text(control, "High"),
        control_rubric_text(control, "Very High"),
    ]


def filled_template_row(control: Control, evaluation: ControlEvaluation) -> List[Any]:
    return template_row(control) + [
        evaluation.security_level,
        evaluation.level_score,
        "Yes" if LEVEL_TO_SCORE[evaluation.security_level] >= LEVEL_TO_SCORE[control.target_security_level] else "No",
        evaluation.priority,
        evaluation.confidence,
        evaluation.answer_summary,
        "\n".join(f"- {item}" for item in evaluation.evidence_used),
        evaluation.gap_summary,
        evaluation.risk_impact,
        "\n".join(f"- {step}" for step in evaluation.remediation_steps),
        evaluation.rationale,
    ]


def assessment_row(control: Control, evaluation: Optional[ControlEvaluation] = None) -> List[Any]:
    if evaluation is None:
        return [
            control.section,
            control.subsection,
            control.control_id,
            control.control_name,
            control.question_text,
            control.evidence_requested,
            control.target_security_level,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            control.weakness_priority,
            "",
            "",
        ]

    return [
        control.section,
        control.subsection,
        control.control_id,
        control.control_name,
        control.question_text,
        control.evidence_requested,
        control.target_security_level,
        evaluation.security_level,
        evaluation.level_score,
        evaluation.answer_summary,
        "\n".join(f"- {item}" for item in evaluation.evidence_used),
        evaluation.gap_summary,
        evaluation.risk_impact,
        "\n".join(f"- {step}" for step in evaluation.remediation_steps),
        evaluation.priority,
        evaluation.confidence,
        evaluation.rationale,
    ]


def write_rubric_sheet(workbook: Workbook, sheet_name: str = "Rubric") -> None:
    rubric_sheet = workbook.create_sheet(sheet_name)
    write_rows(
        rubric_sheet,
        [
            "Security Level",
            "Numeric Score",
            "Meaning",
            "Expected Evidence",
            "Typical Gap",
            "Remediation Focus",
        ],
        RUBRIC_ROWS,
    )


def write_usage_sheet(workbook: Workbook, sheet_name: str = "Instructions") -> None:
    instructions_sheet = workbook.create_sheet(sheet_name)
    instructions_sheet["A1"] = "How To Use This Assessment Framework"
    instructions_sheet["A1"].font = Font(bold=True, size=14)
    instructions_sheet["A3"] = (
        "1. Start from the blank template, the standalone rubric, and the unfilled assessment report template.\n"
        "2. Provide one or more vendor or self-submitted questionnaires to iso_assessment_runner.py.\n"
        "3. Review the filled template and assessment report together.\n"
        "4. Prioritize controls below the target level, especially Critical and High findings.\n"
        "5. Use the remediation guidance as a practical backlog for security improvement."
    )
    instructions_sheet["A5"] = (
        "Scoring Model: each control is rated from Very Low (0) to Very High (4). "
        "Overall scoring is an unweighted maturity average across all controls."
    )
    instructions_sheet["A7"] = (
        "Assessment Coverage: the template includes 256 controls across 16 security domains to evaluate governance, "
        "risk, access control, operations, engineering, suppliers, resilience, privacy, compliance, and supporting evidence."
    )
    instructions_sheet.column_dimensions["A"].width = 120
    instructions_sheet["A3"].alignment = WRAP_ALIGNMENT
    instructions_sheet["A5"].alignment = WRAP_ALIGNMENT
    instructions_sheet["A7"].alignment = WRAP_ALIGNMENT


def control_rubric_text(control: Control, level: str) -> str:
    control_phrase = control.control_name.lower()
    evidence_phrase = control.evidence_requested or "supporting evidence"
    if level == "Very Low":
        return f"No formal or reliable evidence that {control_phrase} is established. {evidence_phrase} is absent or not credible."
    if level == "Low":
        return f"Some activity exists for {control_phrase}, but it appears informal, inconsistent, or narrow in scope, with limited {evidence_phrase.lower()}."
    if level == "Medium":
        return f"{control_phrase.title()} is defined for key areas and partially evidenced, but gaps remain in consistency, review, or coverage."
    if level == "High":
        return f"{control_phrase.title()} is documented, consistently executed, and supported by credible {evidence_phrase.lower()}."
    return f"{control_phrase.title()} is well-governed, measured, and continuously improved, with strong and recurring {evidence_phrase.lower()}."


def write_template_workbook(controls: Sequence[Control], output_path: str) -> None:
    workbook = Workbook()
    template_sheet = workbook.active
    template_sheet.title = "Template"

    template_sheet.append(TEMPLATE_HEADERS)
    for control in controls:
        template_sheet.append(template_row(control))

    style_header_row(template_sheet)
    format_sheet(template_sheet)
    autosize_columns(template_sheet)

    write_rubric_sheet(workbook)
    write_usage_sheet(workbook)

    workbook.save(output_path)


def write_rubric_workbook(output_path: str) -> None:
    workbook = Workbook()
    rubric_sheet = workbook.active
    rubric_sheet.title = "Rubric"
    write_rows(
        rubric_sheet,
        [
            "Security Level",
            "Numeric Score",
            "Meaning",
            "Expected Evidence",
            "Typical Gap",
            "Remediation Focus",
        ],
        RUBRIC_ROWS,
    )
    write_usage_sheet(workbook, "Usage")
    workbook.save(output_path)


def write_filled_template_workbook(
    controls: Sequence[Control],
    evaluations: Dict[str, ControlEvaluation],
    company_name: str,
    output_path: str,
    assessment_mode: str = ASSESSMENT_MODE_QUESTIONNAIRE,
) -> None:
    assessment_mode = normalize_assessment_mode(assessment_mode)
    workbook = Workbook()
    template_sheet = workbook.active
    template_sheet.title = "Template"
    template_sheet.append(FILLED_TEMPLATE_HEADERS)

    for control in controls:
        template_sheet.append(filled_template_row(control, evaluations[control.control_id]))

    style_header_row(template_sheet)
    format_sheet(template_sheet)
    autosize_columns(template_sheet)

    summary_sheet = workbook.create_sheet("Summary")
    metrics = overall_metrics(controls, evaluations)
    summary_sheet["A1"] = "Filled Template Summary"
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet["A3"] = "Company"
    summary_sheet["B3"] = company_name
    summary_sheet["A4"] = "Assessment Mode"
    summary_sheet["B4"] = assessment_mode_label(assessment_mode)
    summary_sheet["A5"] = "Evidence Sufficiency Level"
    summary_sheet["B5"] = metrics["security_level"]
    summary_sheet["A6"] = "Evidence Sufficiency Score Percent"
    summary_sheet["B6"] = metrics["score_percent"]
    summary_sheet["A7"] = "Average Level Score"
    summary_sheet["B7"] = metrics["average_level_score"]
    summary_sheet["A8"] = "Letter Grade"
    summary_sheet["B8"] = metrics["letter_grade"]
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 30

    write_rubric_sheet(workbook)
    write_usage_sheet(workbook)
    workbook.save(output_path)


def write_blank_assessment_workbook(controls: Sequence[Control], output_path: str) -> None:
    workbook = Workbook()
    assessment_sheet = workbook.active
    assessment_sheet.title = "Assessment"
    assessment_sheet.append(ASSESSMENT_HEADERS)
    for control in controls:
        assessment_sheet.append(assessment_row(control))

    style_header_row(assessment_sheet)
    format_sheet(assessment_sheet)
    autosize_columns(assessment_sheet)

    template_sheet = workbook.create_sheet("Template")
    template_sheet.append(TEMPLATE_HEADERS)
    for control in controls:
        template_sheet.append(template_row(control))
    style_header_row(template_sheet)
    format_sheet(template_sheet)
    autosize_columns(template_sheet)

    write_rubric_sheet(workbook)
    write_usage_sheet(workbook)

    workbook.save(output_path)


def write_assessment_workbook(
    controls: Sequence[Control],
    evaluations: Dict[str, ControlEvaluation],
    chunks: Sequence[EvidenceChunk],
    company_name: str,
    model_name: str,
    output_path: str,
    assessment_mode: str = ASSESSMENT_MODE_QUESTIONNAIRE,
) -> Dict[str, Any]:
    assessment_mode = normalize_assessment_mode(assessment_mode)
    metrics = overall_metrics(controls, evaluations)
    section_summary = summarize_sections(controls, evaluations)
    findings = build_findings(controls, evaluations)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Executive Summary"

    strongest_sections = sorted(section_summary, key=lambda row: row["Score Percent"], reverse=True)[:3]
    weakest_sections = section_summary[:3]
    strongest_controls = sorted(
        controls,
        key=lambda control: evaluations[control.control_id].level_score,
        reverse=True,
    )[:5]

    summary_rows = [
        ("Company", company_name),
        ("Assessment Timestamp (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Model", model_name),
        ("Assessment Mode", assessment_mode_label(assessment_mode)),
        ("Evidence Sufficiency Score Percent", metrics["score_percent"]),
        ("Evidence Sufficiency Level", metrics["security_level"]),
        ("Letter Grade", metrics["letter_grade"]),
        ("Level Score Total", metrics["level_score_total"]),
        ("Level Score Max", metrics["level_score_max"]),
        ("Average Level Score", metrics["average_level_score"]),
        ("Controls Assessed", metrics["control_count"]),
        ("Controls Meeting Target", metrics["controls_meeting_target"]),
        ("Controls Below Target", metrics["controls_below_target"]),
        ("Evidence Chunks Loaded", len(chunks)),
    ]

    summary_sheet["A1"] = "Executive Summary"
    summary_sheet["A1"].font = Font(bold=True, size=14)
    summary_sheet["A3"] = "Metric"
    summary_sheet["B3"] = "Value"
    summary_sheet["A3"].fill = HEADER_FILL
    summary_sheet["B3"].fill = HEADER_FILL
    summary_sheet["A3"].font = HEADER_FONT
    summary_sheet["B3"].font = HEADER_FONT

    row_index = 4
    for label, value in summary_rows:
        summary_sheet[f"A{row_index}"] = label
        summary_sheet[f"B{row_index}"] = value
        row_index += 1

    row_index += 1
    summary_sheet[f"A{row_index}"] = "Strongest Sections"
    summary_sheet[f"A{row_index}"].font = SECTION_FONT
    row_index += 1
    for row in strongest_sections:
        summary_sheet[f"A{row_index}"] = row["Section"]
        summary_sheet[f"B{row_index}"] = f"{row['Score Percent']}% ({row['Security Level']})"
        row_index += 1

    row_index += 1
    summary_sheet[f"A{row_index}"] = "Weakest Sections"
    summary_sheet[f"A{row_index}"].font = SECTION_FONT
    row_index += 1
    for row in weakest_sections:
        summary_sheet[f"A{row_index}"] = row["Section"]
        summary_sheet[f"B{row_index}"] = f"{row['Score Percent']}% ({row['Security Level']})"
        row_index += 1

    row_index += 1
    summary_sheet[f"A{row_index}"] = "Top Strength Highlights"
    summary_sheet[f"A{row_index}"].font = SECTION_FONT
    row_index += 1
    for control in strongest_controls:
        evaluation = evaluations[control.control_id]
        summary_sheet[f"A{row_index}"] = f"{control.control_id} - {control.control_name}"
        summary_sheet[f"B{row_index}"] = f"{evaluation.security_level}: {evaluation.answer_summary}"
        row_index += 1

    row_index += 1
    summary_sheet[f"A{row_index}"] = "Highest Priority Findings"
    summary_sheet[f"A{row_index}"].font = SECTION_FONT
    row_index += 1
    for finding in findings[:5]:
        summary_sheet[f"A{row_index}"] = f"{finding['Priority']} | {finding['Control ID']} | {finding['Control Name']}"
        summary_sheet[f"B{row_index}"] = f"{finding['Gap Summary']} Remediation: {finding['Remediation Steps']}"
        row_index += 1

    summary_sheet.column_dimensions["A"].width = 38
    summary_sheet.column_dimensions["B"].width = 120
    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT
            cell.border = THIN_BORDER

    section_sheet = workbook.create_sheet("Section Summary")
    write_rows(
        section_sheet,
        [
            "Section",
            "Control Count",
            "Level Score Total",
            "Level Score Max",
            "Score Percent",
            "Security Level",
            "Below Target Controls",
            "High Priority Findings",
            "Very Low",
            "Low",
            "Medium",
            "High",
            "Very High",
        ],
        section_summary,
    )

    findings_sheet = workbook.create_sheet("Findings")
    write_rows(
        findings_sheet,
        [
            "Priority",
            "Severity Score",
            "Section",
            "Subsection",
            "Control ID",
            "Control Name",
            "Current Level",
            "Target Level",
            "Question",
            "Gap Summary",
            "Risk Impact",
            "Remediation Steps",
            "Evidence Used",
            "Confidence",
        ],
        findings,
    )

    assessment_sheet = workbook.create_sheet("Assessment")
    assessment_sheet.append(ASSESSMENT_HEADERS)

    for control in controls:
        evaluation = evaluations[control.control_id]
        assessment_sheet.append(assessment_row(control, evaluation))

    style_header_row(assessment_sheet)
    format_sheet(assessment_sheet)
    autosize_columns(assessment_sheet)

    for row in range(2, assessment_sheet.max_row + 1):
        level_cell = assessment_sheet[f"H{row}"]
        if level_cell.value in LEVEL_FILLS:
            fill = PatternFill("solid", fgColor=LEVEL_FILLS[level_cell.value])
            for column in ["H", "I"]:
                assessment_sheet[f"{column}{row}"].fill = fill

    template_sheet = workbook.create_sheet("Template")
    template_sheet.append(FILLED_TEMPLATE_HEADERS)
    for control in controls:
        template_sheet.append(filled_template_row(control, evaluations[control.control_id]))
    style_header_row(template_sheet)
    format_sheet(template_sheet)
    autosize_columns(template_sheet)

    write_rubric_sheet(workbook)

    sources_sheet = workbook.create_sheet("Sources")
    grouped_sources = Counter((chunk.document_path, chunk.document_type) for chunk in chunks)
    source_rows = [
        {
            "Source Path": document_path,
            "Source Type": document_type,
            "Excerpt Count": count,
        }
        for (document_path, document_type), count in sorted(grouped_sources.items())
    ]
    write_rows(sources_sheet, ["Source Path", "Source Type", "Excerpt Count"], source_rows)

    workbook.save(output_path)
    return metrics


def validate_runtime_inputs(args: argparse.Namespace) -> None:
    if args.build_only:
        return

    missing = []
    if not args.company:
        missing.append("--company")
    if not args.questionnaires:
        missing.append("--questionnaires")
    if missing:
        joined = ", ".join(missing)
        raise ValueError(f"Missing required arguments for assessment mode: {joined}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Run a detailed, maturity-based ISO-style security assessment using Gemini."
    )
    parser.add_argument("--template", required=True, help="Path to the master control template workbook.")
    parser.add_argument(
        "--assessment_blank",
        help="Optional legacy compatibility path for writing a generated blank workbook.",
    )
    parser.add_argument("--questionnaires", nargs="+", help="One or more questionnaire files to assess.")
    parser.add_argument("--company", help="Company name for the assessment report.")
    parser.add_argument("--output_assessment", default="assessment_report.xlsx")
    parser.add_argument("--model", default="models/gemini-2.5-pro")
    parser.add_argument(
        "--assessment_mode",
        choices=sorted(ASSESSMENT_MODE_LABELS),
        default=ASSESSMENT_MODE_QUESTIONNAIRE,
        help="Prompt mode to use: questionnaire for strict questionnaire review, formal_evidence for artifact-driven review.",
    )
    parser.add_argument("--batch_size", type=int, default=16)
    parser.add_argument(
        "--max_parallel_batches",
        type=int,
        default=1,
        help="Number of independent Gemini control batches to assess at the same time.",
    )
    parser.add_argument("--max_chunks", type=int, default=80)
    parser.add_argument("--top_chunks_per_control", type=int, default=4)
    parser.add_argument("--max_evidence_chars", type=int, default=50000)
    parser.add_argument("--template_output", help="Optional path to write the blank assessment template workbook.")
    parser.add_argument("--rubric_output", help="Optional path to write the standalone rubric workbook.")
    parser.add_argument("--report_template_output", help="Optional path to write the unfilled assessment report workbook.")
    parser.add_argument("--filled_template_output", help="Optional path to write the filled template workbook after assessment.")
    parser.add_argument("--blank_output", help="Deprecated alias for --report_template_output.")
    parser.add_argument(
        "--build_only",
        action="store_true",
        help="Only generate the template, rubric, and assessment report template workbooks.",
    )

    args = parser.parse_args()
    validate_runtime_inputs(args)

    controls = read_template_controls(args.template)

    if args.template_output or args.rubric_output or args.report_template_output or args.blank_output or args.build_only:
        template_output = args.template_output or "security_assessment_template.xlsx"
        rubric_output = args.rubric_output or "security_assessment_rubric.xlsx"
        report_template_output = (
            args.report_template_output
            or args.blank_output
            or args.assessment_blank
            or "assessment_report_template.xlsx"
        )
        write_template_workbook(controls, template_output)
        write_rubric_workbook(rubric_output)
        write_blank_assessment_workbook(controls, report_template_output)
        print(f"[INFO] Wrote template workbook: {template_output}")
        print(f"[INFO] Wrote rubric workbook: {rubric_output}")
        print(f"[INFO] Wrote assessment report template workbook: {report_template_output}")
        if args.build_only:
            return

    api_key, key_source = resolve_api_key_with_source()
    if not api_key:
        raise EnvironmentError(
            "No Gemini API key found. Set GEMINI_API_KEY/GOOGLE_API_KEY or run "
            "'py -3.12 manage_gemini_key.py --set' once to store it locally."
        )
    print(f"[INFO] Using Gemini API key source: {key_source}")

    chunks = load_questionnaire_chunks(args.questionnaires)
    evaluations = assess_controls(
        api_key=api_key,
        model_name=args.model,
        company_name=args.company,
        controls=controls,
        chunks=chunks,
        batch_size=max(1, args.batch_size),
        max_chunks=max(1, args.max_chunks),
        top_chunks_per_control=max(1, args.top_chunks_per_control),
        max_evidence_chars=max(5000, args.max_evidence_chars),
        assessment_mode=args.assessment_mode,
        max_parallel_batches=max(1, args.max_parallel_batches),
    )

    filled_template_output = args.filled_template_output or "security_assessment_template_filled.xlsx"
    write_filled_template_workbook(
        controls=controls,
        evaluations=evaluations,
        company_name=args.company,
        output_path=filled_template_output,
        assessment_mode=args.assessment_mode,
    )

    metrics = write_assessment_workbook(
        controls=controls,
        evaluations=evaluations,
        chunks=chunks,
        company_name=args.company,
        model_name=args.model,
        output_path=args.output_assessment,
        assessment_mode=args.assessment_mode,
    )
    print(f"[INFO] Wrote filled template workbook: {filled_template_output}")
    print(f"[INFO] Wrote assessment report workbook: {args.output_assessment}")

    print(
        f"{args.company}: {metrics['level_score_total']:.2f}/{metrics['level_score_max']:.2f} "
        f"({metrics['score_percent']:.2f}%) | {metrics['security_level']} evidence sufficiency | "
        f"Grade {metrics['letter_grade']}"
    )
